import io
import json
import math
import openpyxl


class ExcelManipulateService :
    def __init__(self, content, sheet_name ="구조",  start_row = 6) :
        self.content = content
        self.wb = openpyxl.load_workbook(io.BytesIO(self.content))
        self.sheet = self.wb[sheet_name]
        self.start_row = start_row
        self.sheet_data = []
        self.target_columns = {
            2: 'B', # megered
            3: 'C', # merged or non-merged
            5: 'E', # unmerged : order
            6: 'F', # unmerged : t&c type
            7: 'G', # unmerged : t&c type name
            8: 'H' # unmerged : isDeleted
        }
        
    def get_all_ranges_by_column(self, column_index, base_range) :
        all_ranges = []

        merged_ranges = self.get_merged_ranges_by_column(column_index) 
        merged_ranges.sort(key=lambda x : x.min_row)
        
        current_row = base_range.min_row
        last_row = base_range.max_row
        
        while current_row <= last_row :
            merged_range = next(
                (r for r in merged_ranges
                if r.min_row <= current_row <= r.max_row),
                None
            )
            
            if merged_range :
                if self.get_cell_value(merged_range.min_row, column_index) is not None :
                    all_ranges.append(merged_range)
                current_row = merged_range.max_row + 1
                
            else :
                from openpyxl.worksheet.merge import MergedCellRange
                single_range = MergedCellRange(
                    self.sheet,
                    f"{openpyxl.utils.get_column_letter(column_index)}{current_row}:"
                    f"{openpyxl.utils.get_column_letter(column_index)}{current_row}"
                )
                if self.get_cell_value(single_range.min_row, column_index) is not None :                  
                    all_ranges.append(single_range)
                current_row += 1
                
        return all_ranges
    
    def get_merged_ranges_by_column(self, column_index) :
        return sorted(
            [range_ for range_ in self.sheet.merged_cells.ranges
            if range_.min_col == column_index and range_.min_row >= self.start_row],
            key = lambda x : x.min_row
        )
        
    def get_cell_value(self, row, column) :
        return self.sheet.cell(row=row, column=column).value
    
    def process_column_data(self, base_range, column_index) :
        column_data = []
        current_row = base_range.min_row
        
        column_merged_ranges = [
            range_ for range_ in self.sheet.merged_cells.ranges
            if range_.min_col == column_index and range_.min_row >= base_range.min_row and range_.max_row <= base_range.max_row
        ]
        column_merged_ranges.sort(key = lambda x : x.min_row)
        
        while current_row <= base_range.max_row:
            merged_range = next(
                (r for r in column_merged_ranges
                if r.min_row <= current_row <= r.max_row),
                None
            )
            
            if merged_range:
                value = self.get_cell_value(merged_range.min_row, column_index)
                if value is not None :
                    column_data.append({
                        'value' : value
                    })
                current_row = merged_range.max_row + 1
            
            else :
                value = self.get_cell_value(current_row, column_index)
                if value is not None :
                    column_data.append({
                        'value' : value
                    })
                current_row += 1
                
        return column_data
    
    def build_basic_data_structure(self) :
        b_merged_ranges = self.get_merged_ranges_by_column(2)
        
        for b_range in b_merged_ranges :
            entry_data = {}
            
            b_value = self.get_cell_value(b_range.min_row, 2)
            entry_data['Country'] = {
                'value' : b_value,
                'terms_lst' : []
            }
            
            c_merged_ranges = self.get_all_ranges_by_column(3, b_range)
            
            for c_range in c_merged_ranges :
                c_value = self.get_cell_value(c_range.min_row, 3)
                
                entry_data['Country']['terms_lst'].append(
                    {
                        c_value : {
                            'tp_code' : []
                        }
                    }
                )
                if self.is_check_merged(c_range) :
                    current_row = c_range.min_row
                    while True :
                        if self.get_cell_value(current_row, 8) != "삭제" :
                            order = self.get_cell_value(current_row, 5)
                            tp_code = self.get_cell_value(current_row, 6)
                            
                            entry_data['Country']['terms_lst'][-1][c_value]['tp_code'].append(tp_code)
                            
                            next_order = self.get_cell_value(current_row+1, 5)
                            if next_order == 1 or next_order is None :
                                break
                            
                            current_row += 1

                else :
                    for c_row in range(c_range.min_row, c_range.max_row + 1) :
                        if self.get_cell_value(c_row, 8) != "삭제" : # allow empty space.
                            entry_data['Country']['terms_lst'][-1][c_value]['tp_code'].append(
                                self.get_cell_value(c_row, 6)
                            )
        
            self.sheet_data.append({'data': entry_data})
            print(f"sheet_data : {self.sheet_data}")
            
    def get_cell_value_for_single_row(self, range, entry_data, c_value) :
        start_row = range.min_row
        nex_cell_value = math.inf
        while(nex_cell_value != 1) :
            cur_cell_value = self.get_cell_value(start_row, 5)
            nex_cell_value = self.get_cell_value(start_row+1, 5)    
            start_row += 1
            
            entry_data['Country']['terms_lst'][-1][c_value]['tp_code'].append(
                self.get_cell_value(start_row, 6)
            )
            
    def is_check_merged(self, range) :
        return bool(range.min_row == range.max_row)

    def get_data(self) :
        ## export json
        json_data = json.dumps(
            self.sheet_data,
            ensure_ascii=False,
            indent=4
        )
        
        with open('eula_data.json', 'w', encoding='utf-8') as f :
            f.write(json_data)
            
        return self.sheet_data