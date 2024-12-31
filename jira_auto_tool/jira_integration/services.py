from http import HTTPStatus
from http.client import HTTPException
from base64 import b64encode
from django.conf import settings
from jira import JIRA, JIRAError
from openpyxl import load_workbook

import io
import json
import os
import httpx
import openpyxl
import pandas as pd
import requests

from dotenv import load_dotenv

class JiraService:
    def __init__(self):
        load_dotenv(verbose=True)
        self.domain = os.getenv("JIRA_SERVER")
        self.username = os.getenv("JIRA_USERNAME")
        self.api_token = os.getenv("JIRA_API_TOKEN")
        self.headers = {
            "Accept" : "application/json",
            "Content-Type" : "application/json"
        }
        
    def connect(self) :
        try :
            self.jira_client = JIRA(
                server=self.domain,
                token_auth=self.api_token
            )
        
            return True
        
        except JIRAError as e :
            # Access error details
            print(f"JIRA Error Status Code: {e.status_code}")
            print(f"JIRA Error Message: {e.text}")
            print(f"JIRA Error URL: {e.url}")
            
            # Handle specific status codes
            if e.status_code == 401:
                print("Authentication failed - check credentials")
            elif e.status_code == 403:
                print("Permission denied")
            elif e.status_code == 404:
                print("Resource not found")
            elif e.status_code == 500:
                print("Internal server error")
            else :
                print("Jira connection fail")
            
            return False
        
    def get_issue(self, issue_key) :
        if not self.jira_client:
            self.connect()
        
        try :
            issue = self.jira_client.issue(issue_key)
        
            
            if hasattr(issue, 'raw') and 'fields' in issue.raw:
                return issue
            
        except Exception as e:
            print(f"Error getting issue: {str(e)}")
            raise
        
    def create_issue(self, project_key, summary, description, issue_type="Story") :
        if not self.jira_client :
            self.connect()
            
        try :
            issue_dict = {
                'project' : project_key,
                'issuetype' : {'name' : issue_type},
                'components' : [{'name' : 'Test'}],
                'summary' : summary,
                'description' : description,
                'assignee' : {'name' : self.username}
            }
            
            issue = self.jira_client.create_issue(fields=issue_dict)
            return issue
        
        except Exception as e :
            print(f"Error creating issue: {str(e)}")
            raise
        
    async def verify_attachment(self, attachment, *description) :
        try :
            platform, npv = description[0], description[1]
            content = attachment.get()
            
            verification_result = {
                'status' : 'verified',
                'checks' : []
            }
            
            # xlsx file verify
            excel_check = self._check_excel_structure(content)
            
            # verification_result['checks'].append(excel_check)
            
            excelReadService = ExcelManipulateService(content)
            excelReadService.build_basic_data_structure()
            
            eula_structure_data = excelReadService.get_data()
            
            res_verify = await self.verify_data_with_api(eula_structure_data, platform, npv) 
            
            verification_result['checks'].append(res_verify)   
                        
            return verification_result
        
        except Exception as e :
            return {
                'status' : 'error',
                'message' : str(e),
                'checks' : []
            }
    
    async def verify_data_with_api(self, data, platform, npv) :
        eula_verificaiton_by_country = {}
        eulaControllerService = EulaControllerService()
        countryControllerService = CountryControllerService()

        eula_data = data
        for eula in eula_data :
            # cntryLst = countryControllerService.process_country_list(eula['data']['Country']['value'])
            termsLst = eula['data']['Country']['terms_lst']

            cleaned_cntry_list = countryControllerService.get_country(eula['data']['Country']['value']) 
            for cntry in cleaned_cntry_list :
                country2Code = countryControllerService.country_mapping.get(cntry, 'Unknown')
                
                if country2Code == 'Unknown' :
                    eula_verificaiton_by_country[cntry] = False
                    continue
                
                apiData = await eulaControllerService.getEulaByCountryAndPlatform(platform, country2Code, npv)


                is_verified = eulaControllerService.compareDataAndSyncStatus(termsLst, apiData)
                country_key = 'global_others' if country2Code == 'JP' else country2Code
                eula_verificaiton_by_country[country_key] = is_verified

        return eula_verificaiton_by_country
    
    def _check_excel_structure(self, content) :
        try :
            wb = load_workbook(io.BytesIO(content), read_only=True)
            
            return {
                'check_type' : 'excel_structure',
                'result' : 'pass',
                'details' : {
                    'worksheet_count' : len(wb.worksheets),
                    'worksheet_names' : wb.sheetnames
                }
            }            
        
        except Exception as e :
            return {
                'check_type' : 'excel_structure',
                'result' : 'fail',
                'details' : {
                    'error' : str(e)
                }
            }
            
class ExcelManipulateService :
    def __init__(self, content, sheet_name ="구조",  start_row = 6) :
        self.content = content
        self.wb = load_workbook(io.BytesIO(self.content))
        self.sheet = self.wb[sheet_name]
        self.start_row = start_row
        self.sheet_data = []
        self.target_columns = {
            2: 'B', # megered
            3: 'C', # merged or non-merged
            5: 'E', # unmerged : order
            6: 'F', # unmerged : t&c type
            7: 'G', # unmerged : t&c type name
            8: 'H' # unmerged : isDeleted
        }
        
    def get_all_ranges_by_column(self, column_index, base_range) :
        all_ranges = []
        
        merged_ranges = self.get_merged_ranges_by_column(column_index)
        merged_ranges.sort(key=lambda x : x.min_row)
        
        current_row = base_range.min_row
        last_row = base_range.max_row
        
        while current_row <= last_row :
            merged_range = next(
                (r for r in merged_ranges
                if r.min_row <= current_row <= r.max_row),
                None
            )
            
            if merged_range :
                if self.get_cell_value(merged_range.min_row, column_index) is not None :
                    all_ranges.append(merged_range)
                current_row = merged_range.max_row + 1
                
            else :
                from openpyxl.worksheet.merge import MergedCellRange
                single_range = MergedCellRange(
                    self.sheet,
                    f"{openpyxl.utils.get_column_letter(column_index)}{current_row}:"
                    f"{openpyxl.utils.get_column_letter(column_index)}{current_row}"
                )
                if self.get_cell_value(single_range.min_row, column_index) is not None :                  
                    all_ranges.append(single_range)
                current_row += 1
                
        return all_ranges
    
    def get_merged_ranges_by_column(self, column_index) :
        return sorted(
            [range_ for range_ in self.sheet.merged_cells.ranges
            if range_.min_col == column_index and range_.min_row >= self.start_row],
            key = lambda x : x.min_row
        )
        
    def get_cell_value(self, row, column) :
        return self.sheet.cell(row=row, column=column).value
    
    def process_column_data(self, base_range, column_index) :
        column_data = []
        current_row = base_range.min_row
        
        column_merged_ranges = [
            range_ for range_ in self.sheet.merged_cells.ranges
            if range_.min_col == column_index and range_.min_row >= base_range.min_row and range_.max_row <= base_range.max_row
        ]
        column_merged_ranges.sort(key = lambda x : x.min_row)
        
        while current_row <= base_range.max_row:
            merged_range = next(
                (r for r in column_merged_ranges
                if r.min_row <= current_row <= r.max_row),
                None
            )
            
            if merged_range:
                value = self.get_cell_value(merged_range.min_row, column_index)
                if value is not None :
                    column_data.append({
                        'value' : value
                    })
                current_row = merged_range.max_row + 1
            
            else :
                value = self.get_cell_value(current_row, column_index)
                if value is not None :
                    column_data.append({
                        'value' : value
                    })
                current_row += 1
                
        return column_data
    
    def build_basic_data_structure(self) :
        b_merged_ranges = self.get_merged_ranges_by_column(2)
        
        for b_range in b_merged_ranges :
            entry_data = {}
            
            b_value = self.get_cell_value(b_range.min_row, 2)
            entry_data['Country'] = {
                'value' : b_value,
                'terms_lst' : []
            }
            
            c_merged_ranges = self.get_all_ranges_by_column(3, b_range)
            
            for c_range in c_merged_ranges :
                c_value = self.get_cell_value(c_range.min_row, 3)
                
                entry_data['Country']['terms_lst'].append(
                    {
                        c_value : {
                            'tp_code' : []
                        }
                    }
                )
                
                for c_row in range(c_range.min_row, c_range.max_row + 1) :
                    if self.get_cell_value(c_row, 8) != "삭제" : # allow empty space.
                        entry_data['Country']['terms_lst'][-1][c_value]['tp_code'].append(
                            self.get_cell_value(c_row, 6)
                        )
        
            self.sheet_data.append({'data': entry_data})

    def get_data(self) :
        ## export json
        json_data = json.dumps(
            self.sheet_data,
            ensure_ascii=False,
            indent=4
        )
        
        with open('eula_data.json', 'w', encoding='utf-8') as f :
            f.write(json_data)
            
        return self.sheet_data
        
class EulaControllerService :
    def __init__(self) :
        self.baseUrl = "http://10.159.73.19:8888/api/v1/terms/terms_group"
        
    async def getEulaByCountryAndPlatform(self, platform: str, country: str, npv: str) :
        async with httpx.AsyncClient() as client :
            try :
                response = await client.get(
                    self.baseUrl,
                    params= {
                        "platform" : platform,
                        "country" : country,
                        "npv" : npv
                    }
                )
                response.raise_for_status()
                
                data = response.json()
                return data
            
            except httpx.TimeoutException :
                return {
                    "error": True,
                    "status_code": HTTPStatus.GATEWAY_TIMEOUT,
                    "detail": {
                        "statusCode": 504,
                        "messages": "Request timeout"
                    }
                }
                
            except httpx.HTTPError as e:
                return {
                    "error": True,
                    "status_code": HTTPStatus.BAD_GATEWAY,
                    "detail": {
                        "statusCode": 502,
                        "messages": f"API error: {str(e)}"
                    }
                }
   
    def compareDataAndSyncStatus(self, data, apidata) :
        responseBody = apidata
        sync_tp = dict()
        sync_tp_result = dict()
        terms_mapping = {
            "음성약관1": "음성 정보",
            "음성약관2": "음성 정보2",
            "LG쇼핑약관": "LG Shopping 약관",
            "맞춤형광고약관": "맞춤형 광고",
            "시청정보약관": "시청 정보",
            "최초동의": "최초 동의",
            "ACR광고약관": "ACR 광고약관",
            "기본약관": "기본 약관",
            "전체동의": "전체 동의"
        }
        
        for d in data :
            terms_name = list(d.keys())[0].strip()
            terms_code = d[list(d.keys())[0]]['tp_code'] 
            terms_name = terms_mapping.get(terms_name, terms_name)
            
            if terms_code != [] :
                sync_tp[terms_name] = terms_code
        
        if 'error' in responseBody and responseBody['error'] :
            pass
        
        try :
            if responseBody['statusCode'] == 200 :
                for tn, tc in sync_tp.items() :
                    if tn in responseBody['response']['terms_lst'].keys() :
                        tpLst = [d['terms_mgt_tp_code'] for d in responseBody['response']['terms_lst'][tn]]
                        sync_tp_result[tn] = set(tpLst) == set(tc)
                            
        except KeyError as e:
            pass
            
        if all(sync_tp_result.values()) :
            return True

        else :
            return False
                
class CountryControllerService :
    def __init__(self) :
        self.country_mapping = {
            # GDPR Countries (18개국)
            'Bulgaria': 'BG',
            'Croatia': 'HR',
            'Cyprus': 'CY',
            'Czech': 'CZ',
            'Estonia': 'EE',
            'Greece': 'GR',
            'Hungary': 'HU',
            'Iceland': 'IS',
            'Latvia': 'LV',
            'Liechtenstein': 'LI',
            'Lithuania': 'LT',
            'Luxemburg': 'LU',
            'Malta': 'MT',
            'Romania': 'RO',
            'Slovakia': 'SK',
            'Slovenia': 'SI',
            'Poland': 'PL',
            'Belgium': 'BE',

            "Switzerland": "CH",
            "Austria": "AT",
            "Finland": "FI",
            "Ireland": "IE",
            "Portugal": "PT",
            "Netherlands": "NL",
            "Sweden": "SE",
            "Denmark": "DK",
            "Norway": "NO",
            
            # GDPR 1개국 (UK)
            'United Kingdom': 'GB',
            'UK': 'GB',
            '영국': 'GB',

            "France": "FR",
            "Spain": "ES",
            "Italy": "IT",
            "Germany": "DE",
            
            "중국" : "CN", # wasu? 

            # Global 12개국
            'United States': 'US',
            'USA': 'US',
            'China': 'CN',

            '뉴질랜드' : 'NZ',
            'New Zealand': 'NZ',
            'Thailand': 'TH',
            'Brazil': 'BR',
            '브라질' : 'BR',
            'Canada': 'CA',
            '캐나다' : 'CA',
            '멕시코' : 'MX',
            'Mexico': 'MX',
            '호주' : 'AU',
            'Australia': 'AU',
            '칠레' : 'CL',
            'Chile': 'CL',
            '아르헨티나' : 'AR',
            'Argentina': 'AR',
            '콜롬비아': 'CO',
            'Colombia': 'CO',
            '페루' : 'PE',
            'Peru': 'PE',
            '인도' : 'IN',
            'India': 'IN',

            # Korea
            'Korea': 'KR',
            '한국': 'KR',
            'South Korea': 'KR',
            
            '태국' : 'TH',
            'Thailand' : 'TH',
            'Kosovo' : 'XK',
            
            # representive Others global
            'Japan' : 'JP'
            }

    def get_country(self, country_value) :
        countryLst = self.process_country_list(country_value)
        countries = countryLst[1:] if len(countryLst) >= 2 else countryLst

        filter_countries = [
            country.strip() for country in countries if country.strip() and not country.startswith("(") and not country.startswith(")") and not country.startswith("*")
        ]

        filter_countries = filter_countries[0].split(', ') if ', ' in filter_countries[0] else filter_countries

        processed_cntry = []
        for country in filter_countries :
            for sub_country in country.split(' / ') :
                processed_cntry.append(sub_country.strip())

        return processed_cntry
    
    def process_country_list(self, country_value) :
        cntryLst = country_value.splitlines()

        if 'Others국가' in cntryLst[0] :
            return ['Japan']
        return cntryLst

        